import openpyxl
from openpyxl.utils import get_column_letter
import json
batch_info_row="1"
batch_columns=[]
batch_no = []
color_code = []
routine = []
total_batch = 0
total_rooms=0
total_slots=0


#name of the file to edit

print("\n____________DIU Routine Batch_no Inserter____________\nAt every step of this script, make sure the file is closed before you proceed.\n")
filename = input("Enter the name of the file. Don't forget to put .xlsx\n>")
deptName = input("\nEnter Department Name: ")
wb = openpyxl.load_workbook(filename)
sheet = wb.active


#counting total number of batches

for column in "DEFGHIJKLMNOPQRSTUVWXYZ":
    if sheet[column+batch_info_row].value is not None:
        total_batch+=1
    else:
        break


#counting total number of slots

col_no=3
while sheet[get_column_letter(col_no)+str(3)].value is not None:
    total_slots+=1
    col_no+=1

batch_columns = list(range(4, 3*total_slots+2, 3))


#counting total number of class rooms

for i in range(4,150):
    if sheet["B"+str(i)].value is not None:
        total_rooms+=1
    else:
        break

total_rooms=total_rooms//6
print("Total Class: "+ str(total_rooms) + "\nTotal batch: "+ str(total_batch)+ "\nSlots: "+ str(total_slots))


#the row that contains batch codes with colors

for i in range(total_batch):
    batch_no.append(sheet[chr(68+i)+"1"].value)
    color_code.append(sheet[chr(68+i)+"1"].fill.fgColor.index)


#inserting data
for row in range(4,4+(total_rooms*6)):
    for col in batch_columns:
        cell = sheet[get_column_letter(col)+str(row)]
        for i in range(len(color_code)):
            if cell.fill.fgColor.index == color_code[i]:
                cell.value = batch_no[i]

wb.save(filename)
print("\nFile Saved\n")


#correcting faulty color codes
while True:
    if (cell_no := input("\nCheck the file again. Any empty cell without batch code? Enter cell no or just hit enter to skip: ").strip()) != '':
        #cell_no=input("Enter the cell no. with faulty color code (Example- D7)")
        color_code=sheet[cell_no].fill.fgColor.index
        batch=int(input("Close file and enter batch no: "))
        for row in range(3,4+(total_rooms*6)):
            for col in batch_columns:
                cell = sheet[get_column_letter(col)+str(row)]
                if cell.fill.fgColor.index == color_code:
                    cell.value = batch
        wb.save(filename)
        print("File Saved")
    else:
        break


#Extrating data from XLSX to an array

for row in range(4,5+(total_rooms*6)):
    day = sheet["A"+str(row)].value
    room = sheet["B"+str(row)].value
    if isinstance(room, float):
        room = int(room)
    for col in batch_columns:
        course = sheet[get_column_letter(col-1) + str(row)].value
        if course:
            teacher = sheet[get_column_letter((col)+1) + str(row)].value
            batch = sheet[get_column_letter(col) + str(row)].value
            if isinstance(batch, float):
                batch = int(batch)
            slot = (col-1)//3
            routine.append([day,slot,str(room),course,teacher,str(batch)])


#Saving the data from array to JSON
filename = deptName+'-routine.json'
open(filename, 'w').write(json.dumps(routine))
print("\nSaved as "+filename)